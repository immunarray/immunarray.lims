# -*- coding: utf-8 -*-
import os
import tempfile
from cStringIO import StringIO

import openpyxl
from Products.Five import BrowserView
from immunarray.lims import normalize
from immunarray.lims.browser.testrun import MissingIChipForSlide
from plone.api.content import find
from plone.api.portal import get


class DownloadXLSX(BrowserView):
    def __init__(self, context, request):
        BrowserView.__init__(self, context, request)
        self.context = context
        self.request = request

    def __call__(self):

        wb = self.load_workbook()
        ws = wb.get_sheet_by_name('Form')

        # A6 - add run number to 'FILE NAME'
        ws['A6'].value += ' (Veracis Test Run %s)' % self.context.run_number
        # A7 - PURPOSE
        ws['A7'].value += ' %s' % self.context.assay_name
        # A8 - DATE
        ws['A8'].value += ' %s' % self.context.run_date
        # D8 - PLANNED BY
        ws['D8'].value += ' %s' % self.context.run_planner
        # H8 - PERFORMED BY
        ws['H8'].value += ' %s' % self.context.run_operator

        # PLATE HEADERS are at 10, 27, 44, 61, 78
        for plate_nr, plate_cell in enumerate([10, 27, 44, 61, 78]):
            plate_nr += 1

            # verify plate header exists in spreadsheet
            plate_title = 'plate %s' % plate_nr
            plate_header_cell = ws['A%s' % plate_cell]
            assert plate_header_cell.value.lower() == plate_title

            # verify the plate exists in run.plates
            if len(self.context.plates) < plate_nr:
                break
            plate = self.context.plates[plate_nr - 1]

            # ALIQUOT WELL ROWS are PLATE HEADER+[2..9] in COLS B,D,F,H
            for chip_nr, chip_col in enumerate('BDFH'):
                for well_nr in range(8):
                    uid = plate['chip-%s_well-%s' % (chip_nr + 1, well_nr + 1)]
                    if uid:
                        brain = find(UID=uid)[0]
                        row = plate_cell + (well_nr + 2)
                        cell = ws['%s%s' % (chip_col, row)]
                        cell.value = brain.Title

            # +1 for ICHIP in rows C,E,G,I  -  and "COMMENTS" in J
            row = plate_cell + 10
            for chip_nr, chip_col in enumerate('CEGI'):
                uid = plate['chip-id-%s' % (chip_nr + 1)]
                if not uid:
                    raise MissingIChipForSlide(
                        "plate %s, slide %s" % (plate_nr, chip_nr + 1))
                brain = find(UID=uid)[0]
                cell = ws['%s%s' % (chip_col, row)]
                cell.value = brain.Title
                # comments in J
                cell = ws['J%s' % row]
                cell.value = plate['comments-ichip-%s' % (chip_nr + 1)]

            # +1 for ICHIP in rows C,E,G,I  -  and "COMMENTS" in J
            row = plate_cell + 10
            for chip_nr, chip_col in enumerate('CEGI'):
                uid = plate['chip-id-%s' % (chip_nr + 1)]
                if not uid:
                    raise MissingIChipForSlide(
                        "plate %s, slide %s" % (plate_nr, chip_nr + 1))
                brain = find(UID=uid)[0]
                cell = ws['%s%s' % (chip_col, row)]
                cell.value = brain.Title
                # comments in J
                cell = ws['J%s' % row]
                cell.value = plate['comments-ichip-%s' % (chip_nr + 1)]

            # +1 for SCAN SLOT C,E,G,I
            row = plate_cell + 11
            for chip_nr, slot_col in enumerate('CEGI'):
                cell = ws['%s%s' % (slot_col, row)]
                cell.value = plate['scan-slot-%s' % (chip_nr + 1)]

            # +2 for COMMENTS in B,D,F,H
            row = plate_cell + 13
            for chip_nr, comment_col in enumerate('CEGI'):
                cell = ws['%s%s' % (comment_col, row)]
                cell.value = plate['comments-ichip-%s' % (chip_nr + 1)]

        fn = "{}-{}".format(self.context.assay_name, self.context.run_number)
        fn = normalize(fn) + ".xlsx"

        setheader = self.request.RESPONSE.setHeader
        setheader('Content-Type', 'application/vnd.ms-excel')
        setheader("Content-Disposition", 'attachment;filename="%s"' % fn)

        fn = tempfile.mktemp()
        wb.save(fn)
        data = open(fn).read()

        self.request.RESPONSE.write(data)

    def load_workbook(self):
        io = StringIO()
        # Get the latest version of the form
        portal = get()
        docs = sorted(portal.documents['pp-007-04-test-forms'].objectValues())
        doc = docs[-1]
        io.write(doc.file.data)
        io.seek(0)
        fn = tempfile.mktemp(suffix='.xlsx')
        open(fn, 'wb').write(io.getvalue())
        wb = openpyxl.load_workbook(fn)
        os.unlink(fn)
        return wb

    def is_hqc(self, sample):
        assay = find(UID=self.context.assay_uid)[0].getObject()
        hqc_id = assay.qc_high_choice
        return sample.id.split('-')[0] == hqc_id

    def is_lqc(self, sample):
        assay = find(UID=self.context.assay_uid)[0].getObject()
        lqc_id = assay.qc_low_choice
        return sample.id.split('-')[0] == lqc_id
