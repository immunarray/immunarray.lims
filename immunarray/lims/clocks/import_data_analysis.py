import logging
import os
import tempfile
from datetime import datetime
from os.path import exists, join
from pprint import pformat
from time import time

import openpyxl
import pdfkit
from AccessControl import Unauthorized
from Products.Five import BrowserView
from Products.Five.browser.pagetemplatefile import ViewPageTemplateFile
from bika.lims.interfaces.limsroot import ILIMSRoot
from immunarray.lims.interfaces.aliquot import IAliquot
from immunarray.lims.interfaces.assayrequest import IAssayRequest
from immunarray.lims.interfaces.clinicalaliquot import IClinicalAliquot
from immunarray.lims.interfaces.clinicalsample import IClinicalSample
from immunarray.lims.interfaces.ichip import IiChip
from immunarray.lims.interfaces.qcaliquot import IQCAliquot
from immunarray.lims.interfaces.sample import ISample
from immunarray.lims.interfaces.veracisrunbase import IVeracisRunBase
from pkg_resources import resource_filename
from plone.api.content import create, find, get_state, transition
from plone.protect.interfaces import IDisableCSRFProtection
from zExceptions import BadRequest
from zope.interface import alsoProvides

logger = logging.getLogger('ImportDataAnalysis')

all_cols = [chr(x) for x in range(65, 91)]  # A..Z


class SpreadsheetParseError(Exception):
    """Error getting required values from spreadsheet.  The parser
    is quite brittle, if anything is unexpected we will probably end up here.
    """


class RunInIncorrectState(Exception):
    """Test run is in the incorrect state.  Test run must be in state 'scanning'
    before results can be imported, and state changed to 'resulted' here.
    """


class ImportDataAnalysis(BrowserView):
    """This runner requires modification of the buildout. Example:

    ```buildout.cfg
        [buildout]
        environment-vars +=
            DATA_ANALYSIS_PATH /path/to/file/drop/folder
            DATA_ANALYSIS_AGE_THRESHOLD 120
            TESTRUN_RESULTS_OUTPUT_PATH /stick/final/pdfs/here

        parts =
            import_data_analysis

        [import_data_analysis]
        <= client_base
        recipe = plone.recipe.zope2instance
        zeo-address = ${zeoserver:zeo-address}
        http-address = <some unused port number>
        zope-conf-additional =
            <clock-server>
               method /Plone/import_data_analysis
               period 60
               user admin
               password adminsecret
               host localhost
            </clock-server>
    ```

    The script can then be invoked with `bin/import_data_analysis start`,
    like any other zeo client, and it will execute the associated method
    every <period> minutes.

    Environment variables are:

    - DATA_ANALYSIS_PATH:
        The full path to the location in which folder containing data analysis
        outputs will be dropped.

    - DATA_ANALYSIS_AGE_THRESHOLD:
        Prevent action while data analysis folders are being uploaded.
        If the current time is not at least this many seconds greater than
        the most recent mtime of all files in the data analysis folder, no
        action is taken.
    
    """

    def __init__(self, context, request):
        alsoProvides(request, IDisableCSRFProtection)
        BrowserView.__init__(self, context, request)
        self.context = context
        self.request = request
        self.filepath = os.environ.get('DATA_ANALYSIS_PATH')
        self.outpath = os.environ.get('TESTRUN_RESULTS_OUTPUT_PATH')
        self.threshold = int(
            os.environ.get('DATA_ANALYSIS_AGE_THRESHOLD', 12000))

    def __call__(self):
        if self.locked():
            return
        try:
            self.lock()
            map(self.process_results, self.get_pending_folders())
        finally:
            self.unlock()

    def locked(self):
        pid = str(os.getpid())
        fn = join(self.filepath, 'lock')
        if exists(fn) and pid in open(fn).read():
            return True
        else:
            self.unlock()
            return False

    def lock(self):
        fn = join(self.filepath, 'lock')
        open(fn, 'w').write(str(os.getpid()))

    def unlock(self):
        fn = join(self.filepath, 'lock')
        if exists(fn):
            os.unlink(fn)

    def get_pending_folders(self):
        """Return a list of full paths, who's file mtimes are outisde
        the safety threshold.
        """
        paths = []
        for fn in os.listdir(self.filepath):
            path = join(self.filepath, fn)
            if not exists(join(path, 'Out', 'Results', 'SLEkey_Results.xlsx')):
                continue
            # Check the modified time
            mtime = os.stat(path).st_mtime
            if time() < mtime + self.threshold:
                logger.info("%s: Newer than threshold." % path)
                continue
            paths.append(path)
        return paths

    def process_results(self, path):
        self.path = path
        self.run = self.get_run(path)
        if not self.run:
            return
        # import results and files from path
        log = self.process_results_xlsx(path)
        log += self.process_out_figures(path)
        log += self.process_raw_flipped(path)
        # transition run to 'resulted'
        self.run.import_log = log
        transition(self.run, 'result')

        # send result report
        self.create_pdf_files()

        # remove uploaded directory
        # rmtree(path)

    def process_results_xlsx(self, path):
        """Take values from path/Out/Results/SLEkey_Results.xlsx, and write
        them to the database
        """
        run = self.get_run(path)

        # Get the first sheet
        filename = join(path, 'Out', 'Results', 'SLEkey_Results.xlsx')
        wb = openpyxl.load_workbook(filename)
        ws = wb.get_sheet_by_name('Sheet1')

        # header rows for blocks containg QC and Clinical aliquot/sample info
        qc_head_row = self.findnextSerial_Numberrow(ws, 1)
        clin_head_row = self.findnextSerial_Numberrow(ws, qc_head_row + 1)

        # get column letters of all qc headers that start with "ichip*"
        qc_ichip_cols = []
        passfail_col = None
        for col in all_cols:
            coord = "%s%s" % (col, qc_head_row)
            val = "%s" % ws[coord].value
            if val:
                if val.lower().startswith("fina"):
                    passfail_col = coord
                if val.lower().startswith("ichip"):
                    qc_ichip_cols.append(col)

        ichip_passfail_combos = []
        if passfail_col:
            for row in range(qc_head_row + 1, clin_head_row):
                passfail = "%s" % ws["%s%s" % (passfail_col, row)].value
                ichip_titles = []
                for col in qc_ichip_cols:
                    ichip_titles.append(ws["%s%s" % (col, row)].value)
                ichip_passfail_combos.append([ichip_titles, passfail])

        # for each plate
        for plate in run.plates:
            plate_ichips = []
            plate_ichiplot_titles = []

            # get all unique ichip lot numbers in a list
            for key, uid in plate.items():
                if key.startswith('ichip-id'):
                    ichip = find(UID=uid)[0].getObject()
                    plate_ichips.append(ichip)
                    ichiplot_title = ichip.title.upper().split('-')[0]
                    if ichiplot_title not in plate_ichiplot_titles:
                        plate_ichiplot_titles.append(ichiplot_title)
            for combo in ichip_passfail_combos:
                if sorted(plate_ichiplot_titles) == sorted(combo[0]):
                    for ichip in plate_ichips:
                        transition(ichip, 'qc_' + combo[1].lower())

            # 'consume' qc aliquots.
            for uid in plate.values():
                brains = find(object_provides=IQCAliquot.__identifier__,
                              UID=uid, review_state='in_process')
                if brains:
                    transition(brains[0].getObject(), 'done')

        # find the column at which the following headers are located:
        cols = {'Sample_ID': '',
                'SLE_key_Score': '',
                'SLE_key_Classification': '',
                'Assay_QC_Status': ''}
        if not all(cols):
            msg = "One of the column headers can't be located:" % pformat(cols)
            raise SpreadsheetParseError(msg)
        for col in all_cols:
            coord = "%s%s" % (col, clin_head_row)
            value = ws[coord].value
            if value in cols:
                cols[value] = col

        # get values from each row and colum into aliquot_results. results is:
        # {sample_id: [SLE_key_Score,
        #              SLE_key_Classification,
        #              Assay_QC_Status],}
        aliquot_results = {}
        first_empty = self.get_first_empty_row(ws, clin_head_row)
        for row in range(clin_head_row + 1, first_empty):
            sample_id = ws["%s%s" % (cols['Sample_ID'], row)].value
            score = ws["%s%s" % (cols['SLE_key_Score'], row)].value
            classif = ws["%s%s" % (cols['SLE_key_Classification'], row)].value
            status = ws["%s%s" % (cols['Assay_QC_Status'], row)].value
            aliquot_results[sample_id] = {
                'SLE_key_Score': score,
                'SLE_key_Classification': "%s" % classif,
                'Assay_QC_Status': "%s" % status
            }

        log = []

        _used_uids = []
        for plate in run.plates:
            for uid in plate.values():
                if uid in _used_uids:
                    continue
                _used_uids.append(uid)
                brains = find(UID=uid)
                if not brains:
                    continue
                aliquot = brains[0].getObject()
                if not IAliquot.providedBy(aliquot):
                    continue
                sample = self.get_sample_from_aliquot(aliquot)
                if IClinicalSample.providedBy(sample):
                    if sample.title not in aliquot_results:
                        msg = "Sample '%s' not in spreadsheet results: %s" % \
                              (sample.title, pformat(aliquot_results))
                        raise RuntimeError(msg)
                    result = aliquot_results[sample.title]
                    ar = self.get_ar_from_sample(sample, run.assay_name)
                    state = result['Assay_QC_Status'].lower()
                    ar.aliquot_evaluated = aliquot.title
                    ar.date_resulted = self.get_date(ws)
                    # assayrequest is associated with multiple aliquots,
                    # so we only transition if the AR's state is expected.
                    if get_state(ar) == 'in_process':
                        transition(ar, 'qc_' + state)
                aliquot.numeric_result = result['SLE_key_Score']
                aliquot.text_result = result['SLE_key_Classification']
                ts = datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M")
                log.append("%s: Sample: %s, Aliquot: %s, Results: %s, %s" % (
                    ts,
                    sample.title,
                    aliquot.title,
                    result['SLE_key_Score'],
                    result['SLE_key_Classification'],
                ))
        return log

    def process_out_figures(self, path):
        """Store images in samples
        """
        figpath = join(path, 'Out', 'Figures')
        log = []
        for fn in os.listdir(figpath):
            if 'png' not in fn:
                continue

            aliquot = self.get_aliquot_from_fn(fn)

            try:
                img = create(container=aliquot, type='Image', id=fn, title=fn)
            except BadRequest as e:
                msg = "Run import has already been performed! (%s)" % e.message
                raise BadRequest(msg)
            except Unauthorized:
                msg = "Failed to create %s in aliquot %s" % (fn, aliquot.title)
                raise Unauthorized(msg)
            ts = datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M")
            log.append("%s: Added image to aliquot %s: %s" %
                       (ts, aliquot.title, img))
        return log

    def create_pdf_files(self):
        path = resource_filename('immunarray.lims', 'reports')
        reports = [r for r in os.listdir(path)
                   if exists(join(path, r, 'report.pt'))]
        # generate report for all aliquots if ANY images exist for them.
        # the report is responsible for assuming correctly which images exist.
        for aliquot in self.get_all_aliquots():
            if not any([self.get_UnivarFigure(aliquot, 1),
                        self.get_UnivarFigure(aliquot, 2),
                        self.get_ReportFigure(aliquot)]):
                continue
            # available as `view.aliquot` in the report's template
            self.aliquot = aliquot
            for report in reports:
              while 1:
                # header and footer must be rendered to actual html files
                options = {}
                htfn = join(path, report, 'report-header.pt')
                if exists(htfn):
                    hfn = tempfile.mktemp(suffix=".html")
                    html = ViewPageTemplateFile(htfn)(self).encode('utf-8')
                    open(hfn, 'w').write(html)
                    options['--header-html'] = hfn
                ftfn = join(path, report, 'report-footer.pt')
                if exists(ftfn):
                    ftn = tempfile.mktemp(suffix=".html")
                    html = ViewPageTemplateFile(ftfn)(self).encode('utf-8')
                    open(ftn, 'w').write(html)
                    options['--footer-html'] = ftn
                # then render the report template itself
                template = ViewPageTemplateFile(join(path, report, 'report.pt'))
                fn = join(self.outpath, '%s - %s' % (aliquot.title, report))
                html = unicode(template(self))

                fn = fn.replace(' ', '')  # XXX debug stuff
                pdfkit.from_string(html, fn + '.pdf', options=options)
                os.system('evince %s' % fn + '.pdf')
                import pdb
                pdb.set_trace()
                pass

    def get_aliquot_from_fn(self, fn):
        sample_id = fn.split('_')[0]
        brains = find(object_provides=ISample.__identifier__, id=sample_id)
        if not brains:
            msg = "%s: Can't find sample '%s'" % (fn, sample_id)
            raise RuntimeError(msg)
        sample = brains[0].getObject()
        # run through the plates, till we find one containing an aliquot
        # which is a child of this sample:
        aliquot = None
        for plate in self.run.plates:
            for key, value in plate.items():
                try:
                    obj = find(UID=value)[0].getObject()
                except IndexError:
                    continue
                if obj.aq_parent.aq_parent == sample:
                    aliquot = obj
                    break
            if aliquot:
                break
        else:
            msg = "Aliquot not found for file: %s" % fn
            raise RuntimeError(msg)
        return aliquot

    def process_raw_flipped(self, path):
        """Store flipped images in ichips
        """
        flpath = join(path, 'Raw', 'Flipped')
        log = []
        for fn in os.listdir(flpath):
            if 'jpg' not in fn:
                continue
            splits = fn.lower().split('_')
            ichip_id = splits[1].replace('.', '-') + "_%03d" % int(splits[2])
            brains = find(object_provides=IiChip.__identifier__, id=ichip_id)
            if not brains:
                msg = "%s/%s: Can't find ichip '%s'" % (flpath, fn, ichip_id)
                raise RuntimeError(msg)
            ichip = brains[0].getObject()
            try:
                img = create(container=ichip, type='Image', id=fn, title=fn)
            except BadRequest as e:
                msg = "Run import has already been performed! (%s)" % e.message
                raise BadRequest(msg)
            except Unauthorized:
                msg = "Failed to create %s in ichip %s" % (fn, ichip.title)
                raise Unauthorized(msg)
            ts = datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M")
            log.append(u"%s: Added image to ichip %s: %s" % (ts, ichip.id, img))
        return log

    def get_run(self, path):
        run_nr = self.get_run_nr(path)
        # get the run nr so we can loop the plates
        brains = find(object_provides=IVeracisRunBase.__identifier__,
                      run_number=run_nr)
        if not brains:
            msg = "No test run found with run_number=%s." % run_nr
            raise RuntimeError(msg)
        run = brains[0].getObject()

        # verify that run is in correct state for import.
        state = get_state(run)
        if state != 'scanning':
            if state == 'resulted':
                return None
            msg = "Test run %s is in state '%s'.  Expected state: 'scanning'" \
                  % (run.title, state)
            raise RunInIncorrectState(msg)

        return run

    def get_run_nr(self, path):
        try:
            csv = [i for i in os.listdir(path) if i.lower().endswith('.csv')][0]
        except Exception as e:
            msg = "Cannot discover run for %s: (%s)" % (path, e.message)
            raise RuntimeError(msg)
        run_nr = int(csv.split('.')[0])
        return run_nr

    def get_sample_from_aliquot(self, aliquot):
        parent = aliquot.aq_parent
        while not (ISample.providedBy(parent)):
            parent = parent.aq_parent
            if ILIMSRoot.providedBy(parent):
                msg = "Cannot find ISample parent of Aliquot: %s" % aliquot
                raise RuntimeError(msg)
        return parent

    def get_ar_from_sample(self, sample, assay_name):
        for y in sample.objectValues():
            if IAssayRequest.providedBy(y) and y.assay_name == assay_name:
                return y

    def get_first_empty_row(self, ws, start):
        for y in range(start, 999):
            if not ws['A%s' % y].value:
                return y

    def findnextSerial_Numberrow(self, ws, start_row):
        """Dig up the next row who's first column contains 'Serial_Number'
        """
        for i in range(start_row, 999):
            value = "%s" % ws['A%s' % i].value
            if value.lower() == 'serial_number':
                return i
        msg = "Can't find cell 'serial_number' in range(A%s,A999)" % start_row
        raise SpreadsheetParseError(msg)

    def get_date(self, ws):
        for y in range(1, 999):
            headervalue = "%s" % ws['A%s' % y].value
            if headervalue.lower() == 'analysis date':
                xy = 'C%s' % y
                value = ws[xy].value
                if isinstance(value, datetime):
                    return value
                msg = "Check that cell has valid date and Date format: %s" % xy
                raise SpreadsheetParseError(msg)
        msg = "Can't find cell 'analysis date' in range(A1,A999)"
        raise SpreadsheetParseError(msg)

    def state(self):
        """return the current review_state of context
        """
        return get_state(self.context)

    def get_ReportFigure(self, aliquot):
        sample = aliquot.aq_parent
        while not ISample.providedBy(sample):
            sample = sample.aq_parent
        fn = join(self.path, 'Out', 'Figures',
                  '%s_ReportFigure.png' % sample.id)
        if exists(fn):
            return fn

    def get_UnivarFigure(self, aliquot, fignr):
        sample = aliquot.aq_parent
        while not ISample.providedBy(sample):
            sample = sample.aq_parent
        fn = join(self.path, 'Raw', 'Flipped', '%s_UnivarFigure_%s.png' %
                  (sample.id, fignr))
        if exists(fn):
            return fn

    def get_all_aliquots(self):
        return self.get_aliquots(IAliquot.__identifier__)

    def get_clinical_aliquots(self):
        return self.get_aliquots(IClinicalAliquot.__identifier__)

    def get_qc_aliquots(self):
        return self.get_aliquots(IQCAliquot.__identifier__)

    def get_aliquots(self, *interfaces):
        items = set()
        for plate in self.run.plates:
            for key, val in plate.items():
                brains = find(object_provides=interfaces, UID=val)
                if brains:
                    items.add(brains[0].getObject())
        return items

    def get_reports_dir(self):
        return resource_filename('immunarray.lims', 'reports')
